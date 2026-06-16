import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from cli_parser import RunMode
from detail_excel_builder import build_detail_excel_workbook
from email_package_orchestrator import run_build_email_package
from report_context_factory import ReportContext
from summary_docx_builder import build_summary_docx_document
from weekly_report_pipeline import run_weekly_report_pipeline


class BuildersIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.project_root = Path(__file__).resolve().parents[1]
        self.output_dir = self.project_root / "tests" / "tmp_output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.excel_path = self.output_dir / "new_source_contract.xlsx"
        self.context = ReportContext(
            mode=RunMode.PREVIEW,
            report_date=date(2026, 3, 30),
            current_week_start=date(2026, 3, 30),
            current_week_end=date(2026, 4, 5),
            next_two_weeks_start=date(2026, 4, 6),
            next_two_weeks_end=date(2026, 4, 19),
            run_id="preview:2026-03-30",
        )
        self._write_source_workbook()

    def _write_source_workbook(self) -> None:
        workbook = Workbook()

        kp_sheet = workbook.active
        kp_sheet.title = "запросы 2026"
        kp_sheet.append(
            [
                "№ КП",
                "Дата получения",
                "Производитель",
                "Филиал",
                "Сотрудник",
                "Тема письма",
                "Крайний срок предоставления предложения",
                "Фактическая дата отправки предложения",
                "Статус подготовки КП",
            ]
        )
        kp_sheet.append(
            [
                1,
                date(2026, 3, 1),
                "Krones",
                "Клин",
                "Катя",
                "КП 1",
                date(2026, 3, 10),
                None,
                "",
            ]
        )
        kp_sheet.append(
            [
                2,
                date(2026, 3, 20),
                "Siemens",
                "Омск",
                "Ира",
                "КП 2",
                date(2026, 3, 27),
                None,
                "",
            ]
        )
        kp_sheet.append(
            [
                3,
                date(2026, 3, 20),
                "KHS",
                "Клин",
                "Катя",
                "КП закрыто",
                date(2026, 3, 25),
                date(2026, 3, 24),
                "отправлено",
            ]
        )
        kp_sheet.append(
            [
                4,
                date(2026, 3, 20),
                "Bad Date",
                "Клин",
                "Катя",
                "Битая строка",
                "31/13/2026",
                None,
                "",
            ]
        )

        delivery_sheet = workbook.create_sheet("заказы 2026")
        delivery_sheet.append(
            [
                None,
                "дата получения заказа",
                "номер РО",
                "Филиал",
                "сумма, евро без НДС",
                "Договор",
                "Ответственный",
                "Дата поставки",
                "Дата отгрузки факт",
            ]
        )
        delivery_sheet.append(
            [
                6187,
                date(2026, 3, 1),
                4501,
                "Клин",
                "7.837,20",
                "C-1",
                "Катя",
                date(2026, 3, 31),
                None,
            ]
        )
        delivery_sheet.append(
            [
                7000,
                date(2026, 3, 5),
                4502,
                "Омск",
                100.0,
                "C-2",
                "Ира",
                date(2026, 4, 10),
                None,
            ]
        )

        workbook.save(self.excel_path)

    def test_pipeline_contains_new_management_sections(self) -> None:
        report = run_weekly_report_pipeline(self.excel_path, self.context)

        self.assertGreater(len(report.manager_summaries), 0)
        self.assertGreater(len(report.manufacturer_bottlenecks), 0)
        self.assertGreater(len(report.branch_bottlenecks), 0)
        self.assertTrue(report.manager_summaries[0].display_name.startswith(("🔴", "🟡", "🟢")))
        self.assertEqual(len(report.source_row_issues), 1)

    def test_detail_excel_workbook_contains_new_sheets_and_headers(self) -> None:
        report = run_weekly_report_pipeline(self.excel_path, self.context)
        workbook = build_detail_excel_workbook(report)

        self.assertIn("Менеджеры", workbook.sheetnames)
        self.assertIn("Узкие места — производители", workbook.sheetnames)
        self.assertIn("Узкие места — филиалы", workbook.sheetnames)
        self.assertIn("🔴", str(workbook["Менеджеры"]["A2"].value))
        self.assertIn("Состояние запроса", str(workbook["Критичные КП"]["J1"].value))

    def test_summary_docx_uses_times_new_roman_black_headings(self) -> None:
        report = run_weekly_report_pipeline(self.excel_path, self.context)
        document = build_summary_docx_document(report)

        heading_style = document.styles["Heading 1"]
        self.assertEqual(heading_style.font.name, "Times New Roman")
        self.assertEqual(heading_style.font.color.rgb, document.styles["Normal"].font.color.rgb)
        self.assertTrue(
            any("Строки, исключенные из расчета" in paragraph.text for paragraph in document.paragraphs)
        )

    def test_email_package_contains_metrics_reference_docx(self) -> None:
        package = run_build_email_package(
            excel_path=self.excel_path,
            context=self.context,
            output_dir=self.output_dir,
        )

        self.assertTrue(package.metrics_reference_docx_path.exists())
        self.assertIn("metrics_reference", package.metrics_reference_docx_path.name)


if __name__ == "__main__":
    unittest.main()
