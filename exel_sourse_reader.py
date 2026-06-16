from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelSourceError(Exception):
    pass


@dataclass(frozen=True)
class RawSheets:
    kp_sheet: Worksheet
    delivery_sheet: Worksheet


def load_excel_workbook(excel_path: Path) -> Workbook:
    if not excel_path.exists():
        raise ExcelSourceError(f"Excel file was not found: {excel_path}")

    try:
        return load_workbook(excel_path)
    except Exception as exc:
        raise ExcelSourceError(f"Could not open Excel file: {excel_path}") from exc


def get_required_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in workbook.sheetnames:
        available_sheets = ", ".join(workbook.sheetnames)
        raise ExcelSourceError(
            f"Required Excel sheet was not found: {sheet_name}. "
            f"Available sheets: {available_sheets}"
        )

    return workbook[sheet_name]


def load_required_sheets(excel_path: Path) -> RawSheets:
    workbook = load_excel_workbook(excel_path)

    kp_sheet = get_required_sheet(workbook, "запросы 2026")
    delivery_sheet = get_required_sheet(workbook, "заказы 2026")

    return RawSheets(
        kp_sheet=kp_sheet,
        delivery_sheet=delivery_sheet,
    )
