from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from entity_mapper import DeliveryItem
from kp_analyzer import KpReportItem, build_request_state
from management_insights import BottleneckSummary, ManagerSummary
from report_metric_catalog import (
    BOTTLENECK_SHEET_HEADERS,
    DELIVERY_SHEET_HEADERS,
    KP_SHEET_HEADERS,
    MANAGER_SHEET_HEADERS,
    SUMMARY_ROW_LABELS,
    SUMMARY_SHEET_HEADERS,
)
from weekly_report_models import BuiltWeeklyReport


def format_date_value(value: date | None) -> str:
    if value is None:
        return ""

    return value.strftime("%d.%m.%Y")


def format_float_value(value: float | None) -> str:
    if value is None:
        return ""

    return f"{value:.2f}"


def format_avg_value(value: float | None) -> str:
    if value is None:
        return ""

    return f"{value:.1f}"


def apply_simple_column_widths(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)

        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 52)


def append_summary_sheet(workbook: Workbook, report: BuiltWeeklyReport) -> None:
    sheet = workbook.active
    sheet.title = "Сводка"

    summary = report.summary

    sheet.append(list(SUMMARY_SHEET_HEADERS))
    rows = [
        (SUMMARY_ROW_LABELS[0], format_date_value(summary.report_date)),
        (SUMMARY_ROW_LABELS[1], summary.critical_kp_count),
        (SUMMARY_ROW_LABELS[2], summary.warning_kp_count),
        (SUMMARY_ROW_LABELS[3], summary.current_week_delivery_count),
        (SUMMARY_ROW_LABELS[4], summary.next_two_weeks_delivery_count),
        (SUMMARY_ROW_LABELS[5], len(report.manager_summaries)),
    ]

    for row in rows:
        sheet.append(list(row))

    apply_simple_column_widths(sheet)


def append_manager_summary_sheet(workbook: Workbook, items: list[ManagerSummary]) -> None:
    sheet = workbook.create_sheet("Менеджеры")

    sheet.append(list(MANAGER_SHEET_HEADERS))

    for item in items:
        sheet.append(
            [
                item.display_name,
                item.overdue_kp_count,
                item.critical_kp_count,
                item.warning_kp_count,
                format_avg_value(item.avg_overdue_days),
                "" if item.max_overdue_days is None else item.max_overdue_days,
                item.current_week_delivery_count,
                item.next_two_weeks_delivery_count,
                item.workload_score,
                item.load_status,
                item.efficiency_assessment,
                item.bottleneck_reason,
                item.delayed_manufacturers,
                item.delayed_branches,
            ]
        )

    apply_simple_column_widths(sheet)


def append_bottleneck_sheet(
    workbook: Workbook,
    title: str,
    items: list[BottleneckSummary],
) -> None:
    sheet = workbook.create_sheet(title)

    sheet.append(list(BOTTLENECK_SHEET_HEADERS))

    for item in items:
        sheet.append(
            [
                item.group_type,
                item.group_name,
                item.overdue_kp_count,
                item.critical_kp_count,
                item.warning_kp_count,
                format_avg_value(item.avg_overdue_days),
                "" if item.max_overdue_days is None else item.max_overdue_days,
                item.affected_managers,
                item.diagnostic_comment,
            ]
        )

    apply_simple_column_widths(sheet)


def append_kp_sheet(
    workbook: Workbook,
    title: str,
    items: list[KpReportItem],
    category_label: str,
) -> None:
    sheet = workbook.create_sheet(title)

    sheet.append(list(KP_SHEET_HEADERS))

    for item in items:
        sheet.append(
            [
                item.source_row_index,
                item.request_number,
                format_date_value(item.received_date),
                item.manufacturer,
                item.branch,
                item.employee,
                item.email_subject,
                format_date_value(item.proposal_deadline),
                item.overdue_days,
                item.request_state or build_request_state(item.overdue_days),
                item.preparation_status,
                category_label,
            ]
        )

    apply_simple_column_widths(sheet)


def append_delivery_sheet(
    workbook: Workbook,
    title: str,
    items: list[DeliveryItem],
    category_label: str,
) -> None:
    sheet = workbook.create_sheet(title)

    sheet.append(list(DELIVERY_SHEET_HEADERS))

    for item in items:
        sheet.append(
            [
                item.source_row_index,
                item.code,
                format_date_value(item.order_received_date),
                item.ro_number,
                item.branch,
                format_float_value(item.amount_eur_without_vat),
                item.contract,
                item.owner,
                format_date_value(item.delivery_date),
                format_date_value(item.shipped_actual_date),
                category_label,
            ]
        )

    apply_simple_column_widths(sheet)


def build_detail_excel_workbook(report: BuiltWeeklyReport) -> Workbook:
    workbook = Workbook()

    append_summary_sheet(workbook, report)
    append_manager_summary_sheet(workbook, report.manager_summaries)
    append_bottleneck_sheet(workbook, "Узкие места — производители", report.manufacturer_bottlenecks)
    append_bottleneck_sheet(workbook, "Узкие места — филиалы", report.branch_bottlenecks)

    append_kp_sheet(
        workbook=workbook,
        title="Критичные КП",
        items=report.critical_kp_items,
        category_label="Критическая просрочка",
    )

    append_kp_sheet(
        workbook=workbook,
        title="КП 1-7 дней",
        items=report.warning_kp_items,
        category_label="Просрочка 1-7 дней",
    )

    append_delivery_sheet(
        workbook=workbook,
        title="Поставки текущая неделя",
        items=report.current_week_delivery_items,
        category_label="Текущая неделя",
    )

    append_delivery_sheet(
        workbook=workbook,
        title="Поставки следующие 2 недели",
        items=report.next_two_weeks_delivery_items,
        category_label="Следующие 2 недели",
    )

    return workbook


def build_detail_excel_file(report: BuiltWeeklyReport, output_path: Path) -> Path:
    workbook = build_detail_excel_workbook(report)
    workbook.save(output_path)
    return output_path
